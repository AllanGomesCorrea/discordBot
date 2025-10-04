import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from typing import Dict, List
import tempfile
import os

class ExcelService:
    def __init__(self):
        pass
    
    def generate_expenses_excel(self, expenses_by_person: Dict[str, List[str]]) -> str:
        """
        Gera um arquivo Excel com as despesas organizadas por pessoa.
        
        Args:
            expenses_by_person: Dicionário com pessoa como chave e lista de despesas como valor
            
        Returns:
            Caminho do arquivo Excel gerado
        """
        # Criar workbook e worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Gastos"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Cabeçalho
        ws['A1'] = "Pessoa"
        ws['B1'] = "Valor"
        ws['C1'] = "Descrição"
        
        # Aplicar estilos ao cabeçalho
        for cell in ['A1', 'B1', 'C1']:
            ws[cell].font = header_font
            ws[cell].fill = header_fill
            ws[cell].alignment = center_alignment
        
        # Dados
        row = 2
        for person in sorted(expenses_by_person.keys()):
            for expense_line in expenses_by_person[person]:
                # Parse da linha: "- valor;descrição;pessoa"
                parts = expense_line[2:].split(';')  # Remove "- " e divide por ";"
                if len(parts) >= 3:
                    value = parts[0]
                    description = parts[1]
                    
                    ws[f'A{row}'] = person
                    ws[f'B{row}'] = float(value.replace(',', '.'))
                    ws[f'C{row}'] = description
                    row += 1
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 40
        
        # Criar arquivo temporário
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_path = temp_file.name
        temp_file.close()
        
        # Salvar arquivo
        wb.save(temp_path)
        
        return temp_path
